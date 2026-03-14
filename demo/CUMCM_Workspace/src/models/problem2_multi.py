"""
problem2_multi.py — 问题3 & 问题4
问题3：FY1 投放3枚干扰弹干扰 M1，最大化总遮蔽时长（并集）
问题4：FY1+FY2+FY3 各1枚干扰弹干扰 M1，最大化总遮蔽时长（并集）
"""
import numpy as np
from scipy.optimize import minimize
import openpyxl, os

G=9.8; V_SINK=3.0; R_EFF=10.0; T_EFF=20.0; V_MISSILE=300.0
V_MIN=70.0; V_MAX=140.0; INTERVAL_MIN=1.0
TRUE_TARGET = np.array([0.0, 200.0, 5.0])
MISSILES = {
    'M1': np.array([20000.,0.,2000.]),
    'M2': np.array([19000.,600.,2100.]),
    'M3': np.array([18000.,-600.,1900.]),
}
DRONES = {
    'FY1': np.array([17800.,0.,1800.]),
    'FY2': np.array([12000.,1400.,1400.]),
    'FY3': np.array([6000.,-3000.,700.]),
    'FY4': np.array([11000.,2000.,1800.]),
    'FY5': np.array([13000.,-2000.,1300.]),
}

def missile_pos(m0,t): return m0*(1-V_MISSILE*t/np.linalg.norm(m0))
def missile_arrival(m0): return np.linalg.norm(m0)/V_MISSILE

def compute_coverage(det_pos, t_det, missile_name='M1', dt=0.001):
    m0=MISSILES[missile_name]; t_arr=missile_arrival(m0)
    t0,t1=t_det,min(t_det+T_EFF,t_arr)
    if t1<=t0: return 0.0,[]
    ts=np.arange(t0,t1+dt,dt)
    fac=V_MISSILE*ts/np.linalg.norm(m0)
    M=m0[np.newaxis,:]*(1-fac[:,np.newaxis])
    C=np.tile(det_pos,(len(ts),1)); C[:,2]-=V_SINK*(ts-t_det)
    d=TRUE_TARGET-M; v=C-M
    ds=np.sum(d**2,axis=1)
    s=np.sum(v*d,axis=1)/np.maximum(ds,1e-12)
    foot=M+s[:,np.newaxis]*d
    dist=np.linalg.norm(C-foot,axis=1)
    ok=(dist<=R_EFF)&(s>=0)&(s<=1)
    ivs,in_c,t_in=[],False,None
    for i,f in enumerate(ok):
        if f and not in_c: in_c=True; t_in=ts[i]
        elif not f and in_c: in_c=False; ivs.append((t_in,ts[i]))
    if in_c: ivs.append((t_in,ts[-1]))
    return sum(b-a for a,b in ivs),ivs

def union_coverage(intervals_list):
    """Compute total length of union of time intervals."""
    all_ivs = sorted([iv for ivs in intervals_list for iv in ivs])
    if not all_ivs: return 0.0,[]
    merged=[list(all_ivs[0])]
    for a,b in all_ivs[1:]:
        if a<=merged[-1][1]: merged[-1][1]=max(merged[-1][1],b)
        else: merged.append([a,b])
    return sum(b-a for a,b in merged), merged

def grenade_traj(d0, angle_deg, v_drone, t_rel, tau_det):
    ar=np.radians(angle_deg)
    vel=np.array([v_drone*np.cos(ar),v_drone*np.sin(ar),0.])
    rp=d0+vel*t_rel
    dp=np.array([rp[0]+vel[0]*tau_det, rp[1]+vel[1]*tau_det,
                 rp[2]-0.5*G*tau_det**2])
    return rp,dp,t_rel+tau_det,vel

def physics_grid_search_single(drone_name, missile_name, delta_y_list=[0.,5.,10.,-5.]):
    """Physics-informed search for best single grenade placement."""
    d0=DRONES[drone_name]; m0=MISSILES[missile_name]
    t_arr=missile_arrival(m0)
    best_cov,best_p=0.,None
    for t_pass in np.arange(3.,t_arr-1,0.3):
        for tau_off in np.arange(0.3,min(T_EFF,t_pass),0.3):
            m_at=missile_pos(m0,t_pass)
            det_z=m_at[2]+V_SINK*tau_off
            if det_z>=d0[2]: continue
            tau_det=np.sqrt(2.*(d0[2]-det_z)/G)
            t_det=t_pass-tau_off
            if t_det<0: continue
            t_rel=t_det-tau_det
            if t_rel<0: continue
            for dy in delta_y_list:
                det_x=m_at[0]; det_y=m_at[1]+dy
                tot=t_det; vx=(det_x-d0[0])/tot; vy=(det_y-d0[1])/tot
                v_d=np.sqrt(vx**2+vy**2)
                if not(V_MIN<=v_d<=V_MAX): continue
                ang=np.degrees(np.arctan2(vy,vx))%360
                dp=np.array([det_x,det_y,det_z])
                c,_=compute_coverage(dp,t_det,missile_name,dt=0.002)
                if c>best_cov: best_cov=c; best_p=[ang,v_d,t_rel,tau_det]
    return best_cov,best_p

# ── 问题3：FY1 × 3枚，干扰 M1 ─────────────────────────────────────────────
def neg_coverage_3grenades(params, missile_name='M1'):
    """
    params = [angle, speed, t_rel_1, tau_det_1, dt_12, tau_det_2, dt_23, tau_det_3]
    dt_12 = t_rel_2 - t_rel_1 >= 1  (间隔约束)
    dt_23 = t_rel_3 - t_rel_2 >= 1
    """
    angle,speed,t1,td1,dt12,td2,dt23,td3=params
    if not(V_MIN<=speed<=V_MAX): return 0.
    if t1<0 or td1<0.1: return 0.
    if dt12<INTERVAL_MIN or td2<0.1: return 0.
    if dt23<INTERVAL_MIN or td3<0.1: return 0.
    t2=t1+dt12; t3=t2+dt23
    d0=DRONES['FY1']
    all_ivs=[]
    for tr,tdx in [(t1,td1),(t2,td2),(t3,td3)]:
        rp,dp,tdet,_=grenade_traj(d0,angle,speed,tr,tdx)
        if dp[2]<=0: continue
        c,ivs=compute_coverage(dp,tdet,missile_name,dt=0.002)
        all_ivs.append(ivs)
    total,_=union_coverage(all_ivs)
    return -total

def solve_problem3():
    print("\n"+"="*60)
    print("  问题3：FY1×3枚 干扰 M1")
    print("="*60)
    # Get best single as seed
    bc,bp=physics_grid_search_single('FY1','M1')
    print(f"  单弹最优参考: {bc:.4f}s")
    angle0,speed0,tr0,td0=bp

    # Build multi-grenade seeds: stagger 3 grenades
    # Grenade 2: delayed by INTERVAL_MIN from grenade1
    # Grenade 3: delayed by 2*INTERVAL_MIN
    seeds=[]
    for dt12 in [1.5,2.,3.,5.]:
        for dt23 in [1.5,2.,3.,5.]:
            for td2 in [td0,td0*1.2,td0*0.8]:
                for td3 in [td0,td0*1.2]:
                    seeds.append([angle0,speed0,tr0,td0,dt12,td2,dt23,td3])
    # Also try problem-1 like seed
    seeds.append([180.,120.,0.,3.6,1.5,3.6,1.5,3.6])

    print(f"  多初始点局部搜索（{len(seeds)}个种子）...")
    best_cov,best_p=0.,None
    for s0 in seeds:
        res=minimize(neg_coverage_3grenades,s0,args=('M1',),method='Nelder-Mead',
                     options={'xatol':1e-4,'fatol':1e-4,'maxiter':5000,'disp':False})
        if -res.fun>best_cov:
            best_cov=-res.fun; best_p=res.x

    angle,speed,t1,td1,dt12,td2,dt23,td3=best_p
    t2=t1+dt12; t3=t2+dt23
    d0=DRONES['FY1']
    detail=[]
    all_ivs=[]
    for i,(tr,tdx) in enumerate([(t1,td1),(t2,td2),(t3,td3)],1):
        rp,dp,tdet,_=grenade_traj(d0,angle%360,speed,tr,tdx)
        c,ivs=compute_coverage(dp,tdet,'M1',dt=0.0005)
        all_ivs.append(ivs)
        detail.append({'id':i,'release_pos':rp,'det_pos':dp,'t_det':tdet,'coverage':c})
    total,merged=union_coverage(all_ivs)

    print(f"\n  飞行方向: {angle%360:.2f}°  速度: {speed:.2f} m/s")
    for d in detail:
        print(f"  弹{d['id']}: 投放点=({d['release_pos'][0]:.2f},{d['release_pos'][1]:.2f},{d['release_pos'][2]:.2f})"
              f"  起爆点=({d['det_pos'][0]:.2f},{d['det_pos'][1]:.2f},{d['det_pos'][2]:.2f})"
              f"  t_det={d['t_det']:.3f}s  单弹={d['coverage']:.4f}s")
    print(f"  总遮蔽时长（并集）: {total:.4f} s")
    print(f"  并集区间: {[(f'{a:.3f}',f'{b:.3f}') for a,b in merged]}")
    return {'angle':angle%360,'speed':speed,'grenades':detail,'total_coverage':total,
            'merged_intervals':merged,'params':best_p}

# ── 问题4：FY1+FY2+FY3 各1枚，干扰 M1 ──────────────────────────────────────
def neg_coverage_3drones(params, missile_name='M1'):
    """
    params = [ang1,spd1,trel1,tdet1, ang2,spd2,trel2,tdet2, ang3,spd3,trel3,tdet3]
    """
    drone_names=['FY1','FY2','FY3']
    all_ivs=[]
    for i,dn in enumerate(drone_names):
        p=params[i*4:(i+1)*4]
        ang,spd,trel,tdx=p
        if not(V_MIN<=spd<=V_MAX) or trel<0 or tdx<0.1: return 0.
        _,dp,tdet,_=grenade_traj(DRONES[dn],ang,spd,trel,tdx)
        if dp[2]<=0: return 0.
        _,ivs=compute_coverage(dp,tdet,missile_name,dt=0.002)
        all_ivs.append(ivs)
    total,_=union_coverage(all_ivs)
    return -total

def solve_problem4():
    print("\n"+"="*60)
    print("  问题4：FY1+FY2+FY3 各1枚 干扰 M1")
    print("="*60)
    drone_names=['FY1','FY2','FY3']
    seeds_per_drone=[]
    for dn in drone_names:
        bc,bp=physics_grid_search_single(dn,'M1',delta_y_list=[0.,5.,10.,-5.,-10.])
        print(f"  {dn} 单弹最优: {bc:.4f}s  params={[f'{x:.3f}' for x in bp]}")
        seeds_per_drone.append(bp)

    # Build combined seed
    seed0=[]
    for bp in seeds_per_drone: seed0+=bp
    print(f"  联合优化（Nelder-Mead）...")
    best_cov,best_p=0.,None
    # Multiple restarts with perturbations
    for trial in range(8):
        noise=np.random.default_rng(trial).normal(0,1,12)*np.tile([5,5,0.5,0.5],3)
        s0=np.array(seed0)+noise
        # clamp speed
        for i in range(3): s0[i*4+1]=np.clip(s0[i*4+1],V_MIN,V_MAX)
        res=minimize(neg_coverage_3drones,s0,args=('M1',),method='Nelder-Mead',
                     options={'xatol':1e-4,'fatol':1e-4,'maxiter':8000,'disp':False})
        if -res.fun>best_cov: best_cov=-res.fun; best_p=res.x

    # Evaluate
    detail=[]
    all_ivs=[]
    for i,dn in enumerate(drone_names):
        ang,spd,trel,tdx=best_p[i*4:(i+1)*4]
        rp,dp,tdet,_=grenade_traj(DRONES[dn],ang%360,spd,trel,tdx)
        c,ivs=compute_coverage(dp,tdet,'M1',dt=0.0005)
        all_ivs.append(ivs)
        detail.append({'drone':dn,'angle':ang%360,'speed':spd,'release_pos':rp,
                       'det_pos':dp,'t_det':tdet,'coverage':c,'intervals':ivs})
    total,merged=union_coverage(all_ivs)

    print(f"\n  ══ 问题4 最优结果 ══")
    for d in detail:
        print(f"  {d['drone']}: 方向={d['angle']:.2f}°  速度={d['speed']:.2f}m/s"
              f"  起爆点=({d['det_pos'][0]:.2f},{d['det_pos'][1]:.2f},{d['det_pos'][2]:.2f})"
              f"  t_det={d['t_det']:.3f}s  单弹={d['coverage']:.4f}s")
    print(f"  总遮蔽时长（并集）: {total:.4f} s")
    print(f"  并集区间: {[(f'{a:.3f}',f'{b:.3f}') for a,b in merged]}")
    return {'drones':detail,'total_coverage':total,'merged_intervals':merged}

# ── 保存 Excel ─────────────────────────────────────────────────────────────
DATA_DIR=os.path.join(os.path.dirname(__file__),'../../data')
OUT_DIR =os.path.join(os.path.dirname(__file__),'../../output')
os.makedirs(OUT_DIR,exist_ok=True)

def save_result1(res3):
    wb=openpyxl.load_workbook(os.path.join(DATA_DIR,'result1.xlsx'))
    ws=wb.active
    angle=res3['angle']; speed=res3['speed']
    for i,g in enumerate(res3['grenades'],2):
        ws.cell(i,1,round(angle,2)); ws.cell(i,2,round(speed,2))
        ws.cell(i,4,round(g['release_pos'][0],3))
        ws.cell(i,5,round(g['release_pos'][1],3))
        ws.cell(i,6,round(g['release_pos'][2],3))
        ws.cell(i,7,round(g['det_pos'][0],3))
        ws.cell(i,8,round(g['det_pos'][1],3))
        ws.cell(i,9,round(g['det_pos'][2],3))
        ws.cell(i,10,round(g['coverage'],4))
    out=os.path.join(OUT_DIR,'result1.xlsx')
    wb.save(out); print(f"  [保存] {out}")

def save_result2(res4):
    wb=openpyxl.load_workbook(os.path.join(DATA_DIR,'result2.xlsx'))
    ws=wb.active
    for i,d in enumerate(res4['drones'],2):
        ws.cell(i,2,round(d['angle'],2)); ws.cell(i,3,round(d['speed'],2))
        ws.cell(i,4,round(d['release_pos'][0],3))
        ws.cell(i,5,round(d['release_pos'][1],3))
        ws.cell(i,6,round(d['release_pos'][2],3))
        ws.cell(i,7,round(d['det_pos'][0],3))
        ws.cell(i,8,round(d['det_pos'][1],3))
        ws.cell(i,9,round(d['det_pos'][2],3))
        ws.cell(i,10,round(d['coverage'],4))
    out=os.path.join(OUT_DIR,'result2.xlsx')
    wb.save(out); print(f"  [保存] {out}")

if __name__=='__main__':
    res3=solve_problem3()
    save_result1(res3)
    res4=solve_problem4()
    save_result2(res4)
    print("\n"+"="*60)
    print(f"  问题3 总遮蔽时长: {res3['total_coverage']:.4f} s")
    print(f"  问题4 总遮蔽时长: {res4['total_coverage']:.4f} s")
    print("="*60)
