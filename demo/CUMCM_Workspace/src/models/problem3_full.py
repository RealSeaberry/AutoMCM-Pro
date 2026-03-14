"""
problem3_full.py — 问题5
5架无人机，各≤3枚干扰弹，干扰 M1+M2+M3
最大化各导弹遮蔽时长之和
"""
import numpy as np
from scipy.optimize import minimize
from itertools import product
import openpyxl, os, json

G=9.8; V_SINK=3.0; R_EFF=10.0; T_EFF=20.0; V_MISSILE=300.0
V_MIN=70.0; V_MAX=140.0; INTERVAL_MIN=1.0
TRUE_TARGET=np.array([0.,200.,5.])
MISSILES={
    'M1':np.array([20000.,0.,2000.]),
    'M2':np.array([19000.,600.,2100.]),
    'M3':np.array([18000.,-600.,1900.]),
}
DRONES={
    'FY1':np.array([17800.,0.,1800.]),
    'FY2':np.array([12000.,1400.,1400.]),
    'FY3':np.array([6000.,-3000.,700.]),
    'FY4':np.array([11000.,2000.,1800.]),
    'FY5':np.array([13000.,-2000.,1300.]),
}

def missile_pos(m0,t): return m0*(1-V_MISSILE*t/np.linalg.norm(m0))
def missile_arrival(m0): return np.linalg.norm(m0)/V_MISSILE

def compute_coverage(det_pos,t_det,missile_name='M1',dt=0.001):
    m0=MISSILES[missile_name]; t_arr=missile_arrival(m0)
    t0,t1=t_det,min(t_det+T_EFF,t_arr)
    if t1<=t0: return 0.,[]
    ts=np.arange(t0,t1+dt,dt)
    fac=V_MISSILE*ts/np.linalg.norm(m0)
    M=m0[np.newaxis,:]*(1-fac[:,np.newaxis])
    C=np.tile(det_pos,(len(ts),1)); C[:,2]-=V_SINK*(ts-t_det)
    d=TRUE_TARGET-M; v=C-M
    ds=np.sum(d**2,axis=1)
    s=np.sum(v*d,axis=1)/np.maximum(ds,1e-12)
    foot=M+s[:,np.newaxis]*d
    dist=np.linalg.norm(C-foot,axis=1)
    ok=(dist<=R_EFF)&(s>=0)&(s<=1)
    ivs,in_c,t_in=[],False,None
    for i,f in enumerate(ok):
        if f and not in_c: in_c=True; t_in=ts[i]
        elif not f and in_c: in_c=False; ivs.append((t_in,ts[i]))
    if in_c: ivs.append((t_in,ts[-1]))
    return sum(b-a for a,b in ivs),ivs

def union_length(ivs_list):
    all_iv=sorted([iv for ivs in ivs_list for iv in ivs])
    if not all_iv: return 0.,[],[]
    merged=[list(all_iv[0])]
    for a,b in all_iv[1:]:
        if a<=merged[-1][1]: merged[-1][1]=max(merged[-1][1],b)
        else: merged.append([a,b])
    return sum(b-a for a,b in merged),merged,all_iv

def grenade_traj(d0,ang,spd,t_rel,tau_det):
    ar=np.radians(ang)
    vel=np.array([spd*np.cos(ar),spd*np.sin(ar),0.])
    rp=d0+vel*t_rel
    dp=np.array([rp[0]+vel[0]*tau_det,rp[1]+vel[1]*tau_det,rp[2]-0.5*G*tau_det**2])
    return rp,dp,t_rel+tau_det,vel

def best_single(drone_name, missile_name,
                delta_y_list=None, t_pass_step=1.0, tau_step=1.0):
    """Physics-driven grid search for best single grenade (fast version)."""
    d0=DRONES[drone_name]; m0=MISSILES[missile_name]
    if delta_y_list is None:
        delta_y_list=[-10.,-5.,0.,5.,10.,15.]
    t_arr=missile_arrival(m0)
    best_c,best_p=0.,None
    for t_pass in np.arange(3.,t_arr-0.5,t_pass_step):
        for tau_off in np.arange(0.5,min(T_EFF,t_pass),tau_step):
            m_at=missile_pos(m0,t_pass)
            det_z=m_at[2]+V_SINK*tau_off
            if det_z>=d0[2]: continue
            tau_det=np.sqrt(2.*(d0[2]-det_z)/G)
            t_det=t_pass-tau_off
            if t_det<0: continue
            t_rel=t_det-tau_det
            if t_rel<0: continue
            for dy in delta_y_list:
                det_x,det_y=m_at[0],m_at[1]+dy
                tot=t_det
                if tot<=0: continue
                vx=(det_x-d0[0])/tot; vy=(det_y-d0[1])/tot
                vd=np.sqrt(vx**2+vy**2)
                if not(V_MIN<=vd<=V_MAX): continue
                ang=np.degrees(np.arctan2(vy,vx))%360
                dp=np.array([det_x,det_y,det_z])
                c,_=compute_coverage(dp,t_det,missile_name,dt=0.003)
                if c>best_c: best_c=c; best_p=[ang,vd,t_rel,tau_det]
    if best_p is None: return 0.,None
    # Local refine
    def neg(p):
        a,v,tr,td=p
        if not(V_MIN<=v<=V_MAX) or tr<0 or td<0.1: return 0.
        _,dp,tdet,_=grenade_traj(d0,a,v,tr,td)
        if dp[2]<=0: return 0.
        c,_=compute_coverage(dp,tdet,missile_name,dt=0.003)
        return -c
    res=minimize(neg,best_p,method='Nelder-Mead',
                 options={'xatol':1e-3,'fatol':1e-3,'maxiter':2000})
    if -res.fun>best_c: best_c=-res.fun; best_p=res.x.tolist()
    return best_c,best_p

# ── 分配策略 ─────────────────────────────────────────────────────────────
# 5架无人机，3枚导弹。每枚导弹至少1架无人机负责。
# 枚举合理分配方案，并对每方案评估总遮蔽时长
DRONE_NAMES=list(DRONES.keys())
MISSILE_NAMES=list(MISSILES.keys())

def evaluate_assignment(assignment, n_grenades_per_drone=None):
    """
    assignment: dict {drone_name: missile_name}  (1对1或1对多)
    n_grenades_per_drone: dict {drone_name: int} (1~3)
    Returns total coverage (sum over missiles of union coverage)
    """
    if n_grenades_per_drone is None:
        n_grenades_per_drone={dn:1 for dn in DRONE_NAMES}

    # Build: missile → list of (drone, n_grenades)
    missile_drones={mn:[] for mn in MISSILE_NAMES}
    for dn,mn in assignment.items():
        if mn is not None:
            missile_drones[mn].append((dn,n_grenades_per_drone.get(dn,1)))

    total=0.
    detail={}
    for mn,drone_list in missile_drones.items():
        if not drone_list:
            detail[mn]={'coverage':0.,'grenades':[]}
            continue
        all_ivs=[]
        grenades=[]
        for dn,ng in drone_list:
            c_best,p_best=best_single(dn,mn)
            if p_best is None: continue
            ang,spd,t_rel,tau_det=p_best
            d0=DRONES[dn]
            rp,dp,tdet,_=grenade_traj(d0,ang%360,spd,t_rel,tau_det)
            c,ivs=compute_coverage(dp,tdet,mn,dt=0.001)
            all_ivs.append(ivs)
            grenades.append({'drone':dn,'n_grenades':ng,'angle':ang%360,
                             'speed':spd,'release_pos':rp,'det_pos':dp,
                             't_det':tdet,'coverage':c,'intervals':ivs})
            # If ng>1: add more grenades with interval constraint
            if ng>1:
                for extra in range(ng-1):
                    new_t_rel=t_rel+(extra+1)*(INTERVAL_MIN+0.5)
                    rp2,dp2,tdet2,_=grenade_traj(d0,ang%360,spd,new_t_rel,tau_det)
                    if dp2[2]<=0: break
                    c2,ivs2=compute_coverage(dp2,tdet2,mn,dt=0.001)
                    all_ivs.append(ivs2)
                    grenades.append({'drone':dn,'n_grenades':0,  # additional
                                     'angle':ang%360,'speed':spd,
                                     'release_pos':rp2,'det_pos':dp2,
                                     't_det':tdet2,'coverage':c2,'intervals':ivs2})
        cov_union,merged,_=union_length(all_ivs)
        total+=cov_union
        detail[mn]={'coverage':cov_union,'merged':merged,'grenades':grenades}
    return total,detail

def solve_problem5():
    print("\n"+"="*60)
    print("  问题5：5架无人机 ≤3枚/架，干扰 M1+M2+M3")
    print("="*60)

    # Step 1: compute best single coverage for each (drone, missile) pair
    print("\n  [1/3] 计算各(无人机,导弹)对的最优单弹遮蔽...")
    cov_table={}  # (drone, missile) -> (coverage, params)
    for dn in DRONE_NAMES:
        for mn in MISSILE_NAMES:
            c,p=best_single(dn,mn)
            cov_table[(dn,mn)]=(c,p)
            print(f"    {dn}→{mn}: {c:.4f}s")

    # Step 2: greedy assignment - assign each drone to best missile
    # Maximize total = sum over missiles of union coverage
    # Simple greedy: for each missile, pick best drone(s)
    print("\n  [2/3] 分配策略搜索（贪心+枚举）...")

    # Try all assignments: each missile gets subset of drones
    # Constraint: total grenades <= 3*5=15, each drone at most 3 grenades
    # Strategy: each drone assigned to exactly 1 missile
    # Enumerate: assign 5 drones to 3 missiles (each missile gets ≥1 drone)

    best_total=0.; best_assign=None; best_detail=None

    # Use pre-computed single values to pick good assignments
    # Sort each drone by its best missile
    drone_pref={}
    for dn in DRONE_NAMES:
        covs=[(cov_table[(dn,mn)][0],mn) for mn in MISSILE_NAMES]
        covs.sort(reverse=True)
        drone_pref[dn]=covs

    # Generate promising assignments by greedy + some perturbations
    from itertools import product as iproduct

    # For each of the 3^5=243 assignments (each drone → one of 3 missiles)
    # filter to those where each missile has ≥1 drone
    count=0
    for assignment_tuple in iproduct(MISSILE_NAMES, repeat=5):
        # Check all missiles covered
        assigned_missiles=set(assignment_tuple)
        if len(assigned_missiles)<3: continue  # must cover all 3 missiles
        count+=1
        assignment={dn:mn for dn,mn in zip(DRONE_NAMES,assignment_tuple)}
        # Simple scoring: sum of best single coverage for each assignment
        score=0.
        for dn,mn in assignment.items():
            score+=cov_table[(dn,mn)][0]
        if score>best_total*0.6:  # prune: only evaluate promising
            # Evaluate with n_grenades based on drone-missile proximity
            ng={}
            for dn,mn in assignment.items():
                c=cov_table[(dn,mn)][0]
                ng[dn]=3 if c>3. else (2 if c>1. else 1)
            total,detail=evaluate_assignment(assignment,ng)
            if total>best_total:
                best_total=total; best_assign=assignment; best_detail=detail

    print(f"  评估了 {count} 种有效分配方案")
    print(f"\n  ══ 问题5 最优分配 ══")
    print(f"  总遮蔽时长（3导弹之和）: {best_total:.4f} s")

    for mn in MISSILE_NAMES:
        d=best_detail[mn]
        print(f"\n  [{mn}] 遮蔽时长: {d['coverage']:.4f}s  并集: {d.get('merged',[])}")
        for g in d['grenades']:
            print(f"    {g['drone']}: 方向={g['angle']:.2f}°  速度={g['speed']:.2f}m/s"
                  f"  起爆=({g['det_pos'][0]:.2f},{g['det_pos'][1]:.2f},{g['det_pos'][2]:.2f})"
                  f"  t_det={g['t_det']:.3f}s  遮蔽={g['coverage']:.4f}s")

    return best_total, best_assign, best_detail

def save_result3(best_assign, best_detail):
    DATA_DIR=os.path.join(os.path.dirname(__file__),'../../data')
    OUT_DIR =os.path.join(os.path.dirname(__file__),'../../output')
    os.makedirs(OUT_DIR,exist_ok=True)
    wb=openpyxl.load_workbook(os.path.join(DATA_DIR,'result3.xlsx'))
    ws=wb.active
    row=2
    # Fill by drone
    for dn in DRONE_NAMES:
        mn=best_assign.get(dn)
        if mn is None: continue
        grenades=[g for g in best_detail.get(mn,{}).get('grenades',[]) if g['drone']==dn]
        for gi,g in enumerate(grenades[:3],1):
            ws.cell(row,1,dn)
            ws.cell(row,2,round(g['angle'],2))
            ws.cell(row,3,round(g['speed'],2))
            ws.cell(row,4,gi)
            ws.cell(row,5,round(g['release_pos'][0],3))
            ws.cell(row,6,round(g['release_pos'][1],3))
            ws.cell(row,7,round(g['release_pos'][2],3))
            ws.cell(row,8,round(g['det_pos'][0],3))
            ws.cell(row,9,round(g['det_pos'][1],3))
            ws.cell(row,10,round(g['det_pos'][2],3))
            ws.cell(row,11,round(g['coverage'],4))
            ws.cell(row,12,mn)
            row+=1
        if not grenades:
            ws.cell(row,1,dn); ws.cell(row,12,mn); row+=1
    out=os.path.join(OUT_DIR,'result3.xlsx')
    wb.save(out); print(f"\n  [保存] {out}")

if __name__=='__main__':
    total5,assign5,detail5=solve_problem5()
    save_result3(assign5,detail5)
    print("\n"+"="*60)
    print(f"  问题5 总遮蔽时长（M1+M2+M3之和）: {total5:.4f} s")
    for mn in MISSILE_NAMES:
        print(f"  {mn}: {detail5[mn]['coverage']:.4f}s")
    print("="*60)
